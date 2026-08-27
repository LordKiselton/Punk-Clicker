# -*- coding: utf-8 -*-
"""Compact game-health snapshot for «ХОЙ!». Writes build/health_latest.md only."""
from __future__ import annotations

import csv
import json
import os
import urllib.parse
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "build"
INBOX = ROOT / "store" / "health_inbox"
APP_ID = "6323422"
LAUNCH = date(2026, 7, 4)
EVENTS = (
    "tutorial_done",
    "daily_claim",
    "ad_watched",
    "prestige",
    "ad_failed",
)

# Rewarded placements (параметр placement у ad_watched / ad_offered / ad_failed).
AD_PLACEMENTS = (
    ("double_gold", "Клад"),
    ("offline_x2", "×2 вход"),
    ("boss_dmg", "×2 урон"),
    ("boss_time", "+15с босс"),
)


def load_token() -> str:
    env = ROOT / ".env"
    if not env.exists():
        raise SystemExit("NO_ENV")
    for line in env.read_text(encoding="utf-8").splitlines():
        if line.strip().startswith("APPMETRICA_OAUTH"):
            return line.split("=", 1)[1].strip().strip('"').strip("'")
    raise SystemExit("NO_TOKEN")


def get_json(token: str, url: str) -> dict | None:
    req = urllib.request.Request(url, headers={"Authorization": f"OAuth {token}"})
    try:
        with urllib.request.urlopen(req, timeout=45) as r:
            return json.loads(r.read().decode("utf-8"))
    except Exception as e:
        return {"_error": str(e)}


def am(token: str, qs: str) -> dict | None:
    url = f"https://api.appmetrica.yandex.ru/stat/v1/data?ids={APP_ID}&{qs}&accuracy=full"
    return get_json(token, url)


def totals(payload: dict | None) -> list[float]:
    if not payload or "_error" in payload or not payload.get("totals"):
        return []
    return [float(x) for x in payload["totals"]]


def event_map(payload: dict | None) -> dict[str, int]:
    out: dict[str, int] = {}
    if not payload or "_error" in payload:
        return out
    for row in payload.get("data") or []:
        name = row["dimensions"][0]["name"]
        out[name] = int(row["metrics"][0])
    return out


def ad_placement_counts(token: str, date1: str, date2: str) -> dict[str, int]:
    """Счётчики ad_watched по placement (просмотры, не уники)."""
    filt = (
        "ym:ce:eventLabel=='ad_watched'"
        " AND ym:ce:paramsLevel1=='placement'"
    )
    qs = (
        f"metrics=ym:ce:clientEvents"
        f"&dimensions=ym:ce:paramsLevel2"
        f"&date1={date1}&date2={date2}"
        f"&filters={urllib.parse.quote(filt)}"
        f"&limit=20&sort=-ym:ce:clientEvents"
    )
    return event_map(am(token, qs))


def fmt_n(x: float | int | None) -> str:
    if x is None:
        return "—"
    if isinstance(x, float) and not x.is_integer():
        return f"{x:.1f}"
    return str(int(x))


def parse_rustore_csvs() -> dict:
    """Two weekly CSVs: larger totals = views, smaller = installs."""
    files = sorted(INBOX.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)[:4]
    parsed = []
    for path in files:
        rows = []
        with path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if not row or row[0].strip().lower().startswith("всего"):
                    continue
                period = row[0].strip()
                try:
                    total = int(float(row[1].replace(" ", "")))
                except (IndexError, ValueError):
                    continue
                rows.append((period, total))
        if not rows:
            continue
        last = rows[-1]
        grand = sum(t for _, t in rows)
        parsed.append({"path": path.name, "last_period": last[0], "last": last[1], "total": grand, "rows": rows})
    if len(parsed) < 2:
        return {"ok": False, "note": "положи 2 CSV RuStore (просмотры + установки) в store/health_inbox/"}
    parsed.sort(key=lambda x: x["total"], reverse=True)
    views, inst = parsed[0], parsed[1]
    cr_all = (100.0 * inst["total"] / views["total"]) if views["total"] else 0
    cr_last = (100.0 * inst["last"] / views["last"]) if views["last"] else 0
    # previous week if present
    cr_prev = None
    if len(views["rows"]) >= 2 and len(inst["rows"]) >= 2:
        pv, pi = views["rows"][-2][1], inst["rows"][-2][1]
        cr_prev = (100.0 * pi / pv) if pv else 0
    return {
        "ok": True,
        "views_file": views["path"],
        "inst_file": inst["path"],
        "period": views["last_period"],
        "views_last": views["last"],
        "inst_last": inst["last"],
        "views_all": views["total"],
        "inst_all": inst["total"],
        "cr_last": cr_last,
        "cr_prev": cr_prev,
        "cr_all": cr_all,
    }


def parse_rsya() -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "note": "openpyxl нет — xlsx не разобран"}
    files = sorted(INBOX.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return {"ok": False, "note": "положи xlsx РСЯ (показы + eCPM) в store/health_inbox/"}
    path = files[0]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def col(name):
        for i, h in enumerate(header):
            if h and name.lower() in str(h).lower():
                return i
        return None
    i_date, i_imp, i_rev, i_ecpm = col("дата"), col("показ"), col("вознагражд"), col("ecpm")
    if i_date is None or i_imp is None or i_rev is None:
        return {"ok": False, "note": f"непонятные колонки в {path.name}: {header}"}
    today = date.today()
    week_ago = today - timedelta(days=6)
    imp = rev = n = 0
    imp7 = rev7 = n7 = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[i_date]
        if not isinstance(d, datetime):
            continue
        shows = float(row[i_imp] or 0)
        money = float(row[i_rev] or 0)
        imp += shows
        rev += money
        n += 1
        if d.date() >= week_ago:
            imp7 += shows
            rev7 += money
            n7 += 1
    ecpm = (1000 * rev / imp) if imp else 0
    ecpm7 = (1000 * rev7 / imp7) if imp7 else 0
    return {
        "ok": True,
        "file": path.name,
        "imp": imp,
        "rev": rev,
        "ecpm": ecpm,
        "imp7": imp7,
        "rev7": rev7,
        "ecpm7": ecpm7,
        "days7": n7,
    }


def main() -> None:
    token = load_token()
    today = date.today()
    d7 = today - timedelta(days=6)
    p7a, p7b = today - timedelta(days=13), today - timedelta(days=7)
    launch, end = LAUNCH.isoformat(), today.isoformat()

    t_all = totals(am(token, f"metrics=ym:u:users,ym:u:newUsers,ym:u:sessions&date1={launch}&date2={end}"))
    t_7 = totals(am(token, f"metrics=ym:u:users,ym:u:newUsers,ym:u:sessions&date1={d7.isoformat()}&date2={end}"))
    t_p = totals(am(token, f"metrics=ym:u:users,ym:u:newUsers,ym:u:sessions&date1={p7a.isoformat()}&date2={p7b.isoformat()}"))
    dur = totals(am(token, f"metrics=ym:s:avgSessionDuration,ym:s:sessions,ym:s:users&date1={d7.isoformat()}&date2={end}"))
    dur_all = totals(am(token, f"metrics=ym:s:avgSessionDuration&date1={launch}&date2={end}"))
    crash = totals(am(token, f"metrics=ym:cr:crashes&date1={launch}&date2={end}"))
    ev_all = event_map(am(token, f"metrics=ym:ce:users&dimensions=ym:ce:eventLabel&date1={launch}&date2={end}&limit=30&sort=-ym:ce:users"))
    ev_7 = event_map(am(token, f"metrics=ym:ce:users&dimensions=ym:ce:eventLabel&date1={d7.isoformat()}&date2={end}&limit=30&sort=-ym:ce:users"))
    ad_all = ad_placement_counts(token, launch, end)
    ad_7 = ad_placement_counts(token, d7.isoformat(), end)
    daily = am(token, f"metrics=ym:u:users,ym:u:newUsers,ym:u:sessions&dimensions=ym:u:date&date1={d7.isoformat()}&date2={end}&limit=14&sort=ym:u:date")

    users, new, sess = (t_all + [None, None, None])[:3]
    u7, n7, s7 = (t_7 + [None, None, None])[:3]
    up, np_, sp = (t_p + [None, None, None])[:3]
    avg7 = dur[0] / 60 if dur else None
    avg_all = dur_all[0] / 60 if dur_all else None
    cr_n = crash[0] if crash else None

    daily_lines = []
    peak = 0
    if daily and daily.get("data"):
        for row in daily["data"]:
            day = row["dimensions"][0]["name"][5:]  # MM-DD
            u, n, s = row["metrics"]
            peak = max(peak, u)
            daily_lines.append(f"| {day} | {int(u)} | {int(n)} | {int(u - n)} | {int(s)} |")

    ru = parse_rustore_csvs()
    ads = parse_rsya()

    def delta(a, b):
        if a is None or b is None or b == 0:
            return "—"
        pct = 100.0 * (a - b) / b
        sign = "+" if pct >= 0 else ""
        return f"{sign}{pct:.0f}%"

    lines = [
        f"# Health snapshot · {today.isoformat()}",
        "",
        f"Окно 7д: **{d7} .. {today}** · vs пред. 7д **{p7a} .. {p7b}** · жизнь с {launch}",
        "",
        "## Store (нужны 2 CSV в inbox)",
    ]
    if ru.get("ok"):
        prev = f"{ru['cr_prev']:.1f}%" if ru.get("cr_prev") is not None else "—"
        lines += [
            f"Последняя неделя RuStore `{ru['period']}`: **{ru['views_last']}** просм. / **{ru['inst_last']}** уст. / CR **{ru['cr_last']:.1f}%** (нед. до: {prev})",
            f"Накопом: {ru['views_all']} просм. / {ru['inst_all']} уст. / CR **{ru['cr_all']:.1f}%**",
        ]
    else:
        lines.append(f"_нет:_ {ru.get('note')}")

    lines += [
        "",
        "## AppMetrica",
        "| | жизнь | 7д | 7д до | WoW |",
        "|--|--:|--:|--:|--:|",
        f"| Уники | {fmt_n(users)} | {fmt_n(u7)} | {fmt_n(up)} | {delta(u7, up)} |",
        f"| Новые | {fmt_n(new)} | {fmt_n(n7)} | {fmt_n(np_)} | {delta(n7, np_)} |",
        f"| Сессии | {fmt_n(sess)} | {fmt_n(s7)} | {fmt_n(sp)} | {delta(s7, sp)} |",
        f"| Сессия, мин | {fmt_n(avg_all) if avg_all else '—'} | {fmt_n(avg7) if avg7 else '—'} | — | — |",
        f"| Краши | {fmt_n(cr_n)} | — | — | — |",
        "",
        "Воронка (уники):",
        "| Событие | жизнь | 7д |",
        "|--|--:|--:|",
    ]
    for ev in EVENTS:
        lines.append(f"| `{ev}` | {ev_all.get(ev, 0)} | {ev_7.get(ev, 0)} |")

    lines += [
        "",
        "Rewarded по placement (`ad_watched`, просмотры):",
        "| Placement | жизнь | 7д | доля жизни |",
        "|--|--:|--:|--:|",
    ]
    ad_sum = sum(ad_all.get(pid, 0) for pid, _ in AD_PLACEMENTS) or sum(ad_all.values())
    for pid, title in AD_PLACEMENTS:
        n = ad_all.get(pid, 0)
        n7 = ad_7.get(pid, 0)
        share = f"{100.0 * n / ad_sum:.0f}%" if ad_sum else "—"
        lines.append(f"| {title} (`{pid}`) | {n} | {n7} | {share} |")
    other = sum(v for k, v in ad_all.items() if k not in {p for p, _ in AD_PLACEMENTS})
    if other:
        lines.append(f"| прочее | {other} | {sum(v for k, v in ad_7.items() if k not in {p for p, _ in AD_PLACEMENTS})} | — |")
    lines.append(f"| **сумма** | **{ad_sum}** | **{sum(ad_7.get(pid, 0) for pid, _ in AD_PLACEMENTS)}** | 100% |")

    if users:
        tut = ev_all.get("tutorial_done", 0)
        lines.append(f"Тутор / уники (жизнь): **{100.0 * tut / users:.0f}%**")
    if n7:
        lines.append(f"Тутор / новые (7д): **{100.0 * ev_7.get('tutorial_done', 0) / n7:.0f}%**")
    if ru.get("ok") and new:
        lines.append(f"Стор-установки накопительно / first open: **{ru['inst_all']} / {int(new)}**")

    lines += ["", "Дни (DAU / new / ret / sess):", "| д | DAU | new | ret | sess |", "|--|--:|--:|--:|--:|"]
    lines += daily_lines or ["| — | — | — | — | — |"]
    if peak:
        lines.append(f"Пик DAU за 7д: **{int(peak)}**")

    lines += ["", "## Ads (xlsx в inbox)"]
    if ads.get("ok"):
        sps = (ads["imp7"] / s7) if (ads.get("imp7") and s7) else None
        lines += [
            f"Файл `{ads['file']}`",
            f"7д: **{int(ads['imp7'])}** показов / **{ads['rev7']:.0f} ₽** / eCPM **{ads['ecpm7']:.0f}**"
            + (f" / {sps:.2f} показа/сессия" if sps else ""),
            f"В файле всего: {int(ads['imp'])} показов / {ads['rev']:.0f} ₽ / eCPM {ads['ecpm']:.0f}",
        ]
    else:
        lines.append(f"_нет:_ {ads.get('note')}")

    lines += [
        "",
        "## Докинь в чат только если есть",
        "- Скрин Retention (D1/D7/D14), если свежий",
        "- Одна фраза: что меняли на карточке на этой неделе",
        "",
        "_Сырой JSON не читать. Оценка — в чате по этому файлу._",
    ]

    OUT.mkdir(exist_ok=True)
    INBOX.mkdir(parents=True, exist_ok=True)
    text = "\n".join(lines) + "\n"
    dest = OUT / "health_latest.md"
    dest.write_text(text, encoding="utf-8")
    print("WROTE", dest)


if __name__ == "__main__":
    main()
